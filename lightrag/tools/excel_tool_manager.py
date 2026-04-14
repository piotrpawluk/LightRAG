"""
Excel Tool Manager — manages Excel-based LLM tools with Redis persistence.

Supports creating, listing, deleting tools, and searching tool data
via full-text search.
"""

import io
import json
from dataclasses import dataclass, asdict
from datetime import datetime, timezone


@dataclass
class ToolParameter:
    column_name: str
    param_name: str
    param_description: str


@dataclass
class ToolDefinition:
    tool_id: str
    name: str
    description: str
    parameters: list[ToolParameter]
    column_names: list[str]
    search_columns: list[str]
    row_count: int
    created_at: str = ""


def parse_excel_file(
    file_data: io.BytesIO, filename: str
) -> tuple[list[str], list[dict], int]:
    """Parse an Excel file and return (columns, rows, row_count).

    Args:
        file_data: File-like object containing Excel data.
        filename: Original filename (used for format validation).

    Returns:
        Tuple of (column_names, rows_as_dicts, row_count).

    Raises:
        ValueError: If file format is invalid or file has no columns.
    """
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise ValueError(
            "Invalid file format. Please upload an .xlsx or .xls file."
        )

    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_data, read_only=True, data_only=True)
        ws = wb.active

        # Extract column names from first row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None or all(c is None for c in header_row):
            raise ValueError("File contains no columns.")

        columns = [str(c) for c in header_row if c is not None]
        if not columns:
            raise ValueError("File contains no columns.")

        # Extract data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, col_name in enumerate(columns):
                val = row[i] if i < len(row) else None
                row_dict[col_name] = val
            rows.append(row_dict)

        wb.close()
        return columns, rows, len(rows)

    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Invalid file format. Please upload an .xlsx or .xls file. Error: {e}")


def generate_tool_schema(tool_def: ToolDefinition) -> dict:
    """Convert a ToolDefinition to an OpenAI-compatible function/tool schema.

    Returns a dict matching OpenAI's tool format:
    {
        "type": "function",
        "function": {
            "name": "...",
            "description": "...",
            "parameters": {
                "type": "object",
                "properties": { ... },
                "required": [ ... ]
            }
        }
    }
    """
    properties = {}
    required = []
    for param in tool_def.parameters:
        properties[param.param_name] = {
            "type": "string",
            "description": param.param_description,
        }
        required.append(param.param_name)

    return {
        "type": "function",
        "function": {
            "name": tool_def.name,
            "description": tool_def.description,
            "parameters": {
                "type": "object",
                "properties": properties,
                "required": required,
            },
        },
    }


class ExcelToolManager:
    """Manages Excel-based LLM tools with Redis persistence and search."""

    def __init__(
        self,
        redis_client,
        namespace: str,
    ):
        self.redis_client = redis_client
        self.namespace = namespace

    def _key(self, *parts: str) -> str:
        return ":".join([self.namespace, "tools", *parts])

    async def create_tool(
        self, tool_def: ToolDefinition, rows: list[dict]
    ) -> ToolDefinition:
        """Create a new tool and persist to Redis.

        Raises:
            ValueError: If a tool with the same name already exists.
        """
        # Check for duplicate name
        name_key = self._key("names")
        if await self.redis_client.sismember(name_key, tool_def.name):
            raise ValueError(f"A tool with the name '{tool_def.name}' already exists.")

        if not tool_def.created_at:
            tool_def.created_at = datetime.now(timezone.utc).isoformat()

        # Persist tool metadata
        meta = {
            "tool_id": tool_def.tool_id,
            "name": tool_def.name,
            "description": tool_def.description,
            "parameters": [asdict(p) for p in tool_def.parameters],
            "column_names": tool_def.column_names,
            "search_columns": tool_def.search_columns,
            "row_count": tool_def.row_count,
            "created_at": tool_def.created_at,
        }
        await self.redis_client.set(
            self._key(tool_def.tool_id, "meta"),
            json.dumps(meta),
        )

        # Persist row data
        data = {
            "rows": rows,
            "search_columns": tool_def.search_columns,
        }
        await self.redis_client.set(
            self._key(tool_def.tool_id, "data"),
            json.dumps(data, default=str),
        )

        # Add to index and name set
        await self.redis_client.sadd(self._key("index"), tool_def.tool_id)
        await self.redis_client.sadd(name_key, tool_def.name)

        return tool_def

    async def list_tools(self) -> list[dict]:
        """List all registered tools."""
        tool_ids = await self.redis_client.smembers(self._key("index"))
        tools = []
        for tid in tool_ids:
            tid_str = tid.decode() if isinstance(tid, bytes) else tid
            meta_json = await self.redis_client.get(self._key(tid_str, "meta"))
            if meta_json:
                meta = json.loads(meta_json)
                tools.append(meta)
        return tools

    async def get_tool(self, tool_id: str) -> dict | None:
        """Get a single tool's metadata."""
        meta_json = await self.redis_client.get(self._key(tool_id, "meta"))
        if meta_json:
            return json.loads(meta_json)
        return None

    async def delete_tool(self, tool_id: str) -> bool:
        """Delete a tool and all its data from Redis."""
        meta = await self.get_tool(tool_id)
        if meta:
            await self.redis_client.srem(self._key("names"), meta["name"])

        await self.redis_client.srem(self._key("index"), tool_id)
        await self.redis_client.delete(self._key(tool_id, "meta"))
        await self.redis_client.delete(self._key(tool_id, "data"))
        return True

    async def search(
        self,
        tool_id: str,
        params: dict[str, str],
        top_k: int = 10,
    ) -> list[dict]:
        """Search tool data using full-text search.

        Args:
            tool_id: The tool to search.
            params: Dict mapping param_name to search value.
            top_k: Maximum rows to return.

        Returns:
            List of matching row dicts.
        """
        # Load tool data
        meta_json = await self.redis_client.get(self._key(tool_id, "meta"))
        if not meta_json:
            return []
        meta = json.loads(meta_json)

        data_json = await self.redis_client.get(self._key(tool_id, "data"))
        if not data_json:
            return []
        data = json.loads(data_json)
        rows = data["rows"]
        search_columns = data.get("search_columns", meta.get("search_columns", []))

        # Build param_name -> column_name mapping from tool meta
        param_to_col = {}
        for p in meta.get("parameters", []):
            param_to_col[p["param_name"]] = p["column_name"]

        # Score each row
        row_scores: list[tuple[int, float]] = []

        for idx, row in enumerate(rows):
            score = 0.0

            for param_name, search_value in params.items():
                col_name = param_to_col.get(param_name, param_name)
                if col_name not in search_columns:
                    continue

                search_lower = str(search_value).lower()
                cell_value = str(row.get(col_name, "")).lower()

                # Full-text match: substring matching
                if search_lower in cell_value:
                    score += 2.0  # Strong boost for exact substring match
                elif any(word in cell_value for word in search_lower.split()):
                    score += 1.0  # Partial word match

            if score > 0:
                row_scores.append((idx, score))

        # Sort by score descending, take top_k
        row_scores.sort(key=lambda x: x[1], reverse=True)
        return [rows[idx] for idx, _ in row_scores[:top_k]]

    async def store_temp_upload(
        self, file_id: str, columns: list[str], rows: list[dict]
    ) -> None:
        """Store uploaded file data temporarily (TTL 30min)."""
        data = {"columns": columns, "rows": rows}
        key = self._key("upload", file_id)
        await self.redis_client.set(key, json.dumps(data, default=str))
        await self.redis_client.expire(key, 1800)  # 30 minutes

    async def get_temp_upload(self, file_id: str) -> dict | None:
        """Retrieve temporarily stored upload data."""
        key = self._key("upload", file_id)
        data_json = await self.redis_client.get(key)
        if data_json:
            return json.loads(data_json)
        return None

    async def get_all_tool_definitions(self) -> list[ToolDefinition]:
        """Load all tools as ToolDefinition objects (for LLM schema generation)."""
        tools_meta = await self.list_tools()
        result = []
        for meta in tools_meta:
            params = [
                ToolParameter(**p) for p in meta.get("parameters", [])
            ]
            td = ToolDefinition(
                tool_id=meta["tool_id"],
                name=meta["name"],
                description=meta["description"],
                parameters=params,
                column_names=meta.get("column_names", []),
                search_columns=meta.get("search_columns", []),
                row_count=meta.get("row_count", 0),
                created_at=meta.get("created_at", ""),
            )
            result.append(td)
        return result
