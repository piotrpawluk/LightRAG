#!/usr/bin/env python3
"""
Create a test Excel file for demonstrating excel_lightrag_query.py

Usage:
    python scripts/create_test_excel.py
"""

import openpyxl
from pathlib import Path


def create_test_file():
    """Create a test Excel file with sample prompts."""
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"

    # Add headers
    ws["A1"] = "ID"
    ws["B1"] = "Pytanie"  # Polish for "Question"
    ws["C1"] = "Category"
    ws["U1"] = "Response"  # This will be filled by the script

    # Add sample prompts
    test_prompts = [
        ("1", "What is LightRAG?", "General"),
        ("2", "How does retrieval-augmented generation work?", "Technical"),
        ("3", "What are the benefits of using a knowledge graph?", "Concepts"),
        ("4", "Explain the difference between local and global query modes", "Features"),
        ("5", "What storage backends does LightRAG support?", "Technical"),
    ]

    for idx, (id_val, question, category) in enumerate(test_prompts, start=2):
        ws[f"A{idx}"] = id_val
        ws[f"B{idx}"] = question
        ws[f"C{idx}"] = category

    # Adjust column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["U"].width = 80

    # Save file
    output_path = Path(__file__).parent / "test_data.xlsx"
    wb.save(output_path)
    print(f"✓ Created test file: {output_path}")
    print(f"  - Contains {len(test_prompts)} sample prompts")
    print(f"  - Prompt column: B (Pytanie)")
    print(f"  - Response column: U (initially empty)")
    print("\nRun the query processor with:")
    print(f"  python scripts/excel_lightrag_query.py --excel-file {output_path} --dry-run")


if __name__ == "__main__":
    create_test_file()
