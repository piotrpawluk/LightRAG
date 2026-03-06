#!/usr/bin/env python3
"""
Multi-Sheet Excel-to-LightRAG Query Processor

This script processes an Excel file by:
1. Iterating through all sheets from the 2nd sheet to the last (skips 1st sheet)
2. Reading prompts from column B (with header row, so starts at row 2)
3. Querying the LightRAG API in parallel (max 3 concurrent requests)
4. Writing responses to column I

Usage:
    python scripts/excel_lightrag_query_multisheet.py --excel-file data.xlsx --api-url http://localhost:9621
"""

import argparse
import asyncio
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import aiohttp
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from tenacity import (
    retry,
    stop_after_attempt,
    wait_exponential,
    retry_if_exception_type,
)

try:
    from tqdm.asyncio import tqdm as async_tqdm
except ImportError:
    print("Warning: tqdm not installed. Install with: pip install tqdm")
    print("Progress bar will not be available.")
    async_tqdm = None


class MultiSheetExcelProcessor:
    """Handles Excel file reading and writing operations across multiple sheets."""

    def __init__(
        self,
        file_path: str,
        prompt_column: str = "B",
        output_column: str = "I",
        start_sheet: int = 2,
    ):
        self.file_path = file_path
        self.workbook = None
        self.PROMPT_COLUMN = prompt_column.upper()
        self.OUTPUT_COLUMN = output_column.upper()
        self.start_sheet = start_sheet  # 1-indexed (human-readable)

    def load(self):
        """Load Excel file and validate it has enough sheets."""
        try:
            self.workbook = load_workbook(self.file_path)
            sheet_count = len(self.workbook.worksheets)

            if sheet_count < self.start_sheet:
                raise RuntimeError(
                    f"Excel file has only {sheet_count} sheet(s), "
                    f"but start_sheet is set to {self.start_sheet}"
                )

            print(f"✓ Loaded Excel file: {self.file_path}")
            print(f"✓ Total sheets: {sheet_count}")
        except Exception as e:
            if "sheet" in str(e).lower():
                raise
            raise RuntimeError(f"Failed to load Excel file: {e}")

    def get_sheets_to_process(self) -> list[Worksheet]:
        """
        Return sheets from start_sheet index onwards.

        Returns:
            List of worksheets to process (start_sheet to last)
        """
        # Convert 1-indexed start_sheet to 0-indexed
        start_index = self.start_sheet - 1
        return self.workbook.worksheets[start_index:]

    def get_prompts(self, worksheet: Worksheet) -> list[tuple[int, str]]:
        """
        Extract all non-empty prompts from the prompt column of a given worksheet.
        Skips the header row (row 1).

        Args:
            worksheet: The worksheet to extract prompts from

        Returns:
            List of tuples (row_number, prompt_text)
        """
        prompts = []
        max_row = worksheet.max_row

        for row_num in range(2, max_row + 1):  # Skip header row
            cell_value = worksheet[f"{self.PROMPT_COLUMN}{row_num}"].value
            if cell_value and str(cell_value).strip():
                prompts.append((row_num, str(cell_value).strip()))

        return prompts

    def write_response(self, worksheet: Worksheet, row_num: int, response: str):
        """
        Write response to the output column for the given row in the specified worksheet.
        Always overwrites existing data.

        Args:
            worksheet: The worksheet to write to
            row_num: Excel row number
            response: Response text to write
        """
        worksheet[f"{self.OUTPUT_COLUMN}{row_num}"] = response

    def save(self, backup: bool = True):
        """
        Save the workbook with optional timestamped backup.

        Args:
            backup: If True, creates a backup before saving
        """
        if backup:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{self.file_path}.backup.{timestamp}"
            shutil.copy2(self.file_path, backup_path)
            print(f"✓ Backup created: {backup_path}")

        self.workbook.save(self.file_path)
        print(f"✓ Results saved to: {self.file_path}")


class LightRAGClient:
    """Async HTTP client for LightRAG API."""

    def __init__(
        self,
        base_url: str,
        api_key: Optional[str] = None,
        token: Optional[str] = None,
        timeout: int = 120,
    ):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.token = token
        self.timeout = aiohttp.ClientTimeout(total=timeout)
        self.session: Optional[aiohttp.ClientSession] = None

    async def __aenter__(self):
        """Create aiohttp session with timeout configuration."""
        self.session = aiohttp.ClientSession(timeout=self.timeout)
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Close aiohttp session."""
        if self.session:
            await self.session.close()

    async def login(self, username: str, password: str) -> str:
        """
        Authenticate with username/password and return JWT token.

        Args:
            username: Username for authentication
            password: Password for authentication

        Returns:
            JWT access token

        Raises:
            aiohttp.ClientError: On HTTP errors
        """
        url = f"{self.base_url}/login"
        data = aiohttp.FormData()
        data.add_field("username", username)
        data.add_field("password", password)

        async with self.session.post(url, data=data) as resp:
            resp.raise_for_status()
            result = await resp.json()
            return result.get("access_token")

    async def query(self, prompt: str) -> dict:
        """
        Send query to LightRAG API.

        Args:
            prompt: Query text

        Returns:
            JSON response from API

        Raises:
            aiohttp.ClientError: On HTTP errors
            asyncio.TimeoutError: On timeout
        """
        url = f"{self.base_url}/query"
        headers = {"Content-Type": "application/json"}

        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        elif self.api_key:
            headers["X-API-Key"] = self.api_key

        payload = {
            "query": prompt,
            "mode": "global",
            "top_k": 40,
            "chunk_top_k": 20,
            "max_entity_tokens": 6000,
            "max_relation_tokens": 8000,
            "max_total_tokens": 30000,
            "enable_rerank": False,
            "include_references": True,
            "response_type": "Single Paragraph"
        }

        async with self.session.post(url, json=payload, headers=headers) as resp:
            resp.raise_for_status()
            return await resp.json()


class ParallelQueryProcessor:
    """Manages parallel query processing with concurrency control."""

    def __init__(self, client: LightRAGClient, max_concurrent: int = 3):
        self.client = client
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.max_concurrent = max_concurrent

    async def process_query(
        self, row_num: int, prompt: str
    ) -> tuple[int, str, bool]:
        """
        Process a single query with retry logic.

        Args:
            row_num: Excel row number
            prompt: Query text

        Returns:
            Tuple of (row_number, response_text, success_flag)
        """
        async with self.semaphore:

            @retry(
                stop=stop_after_attempt(3),
                wait=wait_exponential(multiplier=1, min=2, max=10),
                retry=retry_if_exception_type(
                    (aiohttp.ClientError, asyncio.TimeoutError)
                ),
            )
            async def _query_with_retry():
                result = await self.client.query(prompt)
                return result.get("response", "")

            try:
                response_text = await _query_with_retry()
                return (row_num, response_text, True)
            except Exception as e:
                error_msg = f"ERROR (after 3 retries): {str(e)}"
                return (row_num, error_msg, False)

    async def process_all(
        self, prompts: list[tuple[int, str]], sheet_name: str = ""
    ) -> list[tuple[int, str, bool]]:
        """
        Process all prompts with progress tracking.

        Args:
            prompts: List of (row_number, prompt_text) tuples
            sheet_name: Name of the current sheet (for progress display)

        Returns:
            List of (row_number, response_text, success_flag) tuples
        """
        tasks = [
            self.process_query(row_num, prompt) for row_num, prompt in prompts
        ]

        desc = f"Processing '{sheet_name}'" if sheet_name else "Processing queries"

        if async_tqdm:
            # Use tqdm for detailed progress bar
            results = []
            for coro in async_tqdm.as_completed(
                tasks,
                desc=desc,
                total=len(tasks),
                unit="query",
            ):
                results.append(await coro)
            # Sort by row number to maintain order
            results.sort(key=lambda x: x[0])
        else:
            # Fallback without progress bar
            print(f"{desc}: {len(tasks)} queries (max {self.max_concurrent} concurrent)...")
            results = await asyncio.gather(*tasks)

        return results


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Process multiple Excel sheets with LightRAG queries",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage (processes 2nd sheet onwards, reads from B, writes to I)
  python scripts/excel_lightrag_query_multisheet.py --excel-file data.xlsx

  # With authentication and backup
  python scripts/excel_lightrag_query_multisheet.py --excel-file data.xlsx --api-key YOUR_KEY --backup

  # Start from 3rd sheet instead of 2nd
  python scripts/excel_lightrag_query_multisheet.py --excel-file data.xlsx --start-sheet 3

  # Custom columns
  python scripts/excel_lightrag_query_multisheet.py --excel-file data.xlsx --prompt-column C --output-column J

  # Dry run to validate without writing
  python scripts/excel_lightrag_query_multisheet.py --excel-file data.xlsx --dry-run

Environment Variables:
  LIGHTRAG_API_URL      Default API URL (default: http://localhost:9621)
  LIGHTRAG_API_KEY      API key for authentication
  LIGHTRAG_USERNAME     Username for login authentication
  LIGHTRAG_PASSWORD     Password for login authentication
        """,
    )

    parser.add_argument(
        "--excel-file",
        required=True,
        help="Path to Excel file (.xlsx)",
    )

    parser.add_argument(
        "--api-url",
        default=os.getenv("LIGHTRAG_API_URL", "http://localhost:9621"),
        help="LightRAG API base URL (default: $LIGHTRAG_API_URL or http://localhost:9621)",
    )

    parser.add_argument(
        "--api-key",
        default=os.getenv("LIGHTRAG_API_KEY"),
        help="API key for authentication (default: $LIGHTRAG_API_KEY)",
    )

    parser.add_argument(
        "--username",
        default=os.getenv("LIGHTRAG_USERNAME"),
        help="Username for login authentication (default: $LIGHTRAG_USERNAME)",
    )

    parser.add_argument(
        "--password",
        default=os.getenv("LIGHTRAG_PASSWORD"),
        help="Password for login authentication (default: $LIGHTRAG_PASSWORD)",
    )

    parser.add_argument(
        "--max-concurrent",
        type=int,
        default=3,
        help="Maximum concurrent requests (default: 3)",
    )

    parser.add_argument(
        "--timeout",
        type=int,
        default=120,
        help="Request timeout in seconds (default: 120)",
    )

    parser.add_argument(
        "--output-column",
        default="I",
        help="Column for writing responses (default: I)",
    )

    parser.add_argument(
        "--prompt-column",
        default="B",
        help="Column containing prompts (default: B)",
    )

    parser.add_argument(
        "--start-sheet",
        type=int,
        default=2,
        help="First sheet to process (1-indexed, default: 2 = second sheet)",
    )

    parser.add_argument(
        "--backup",
        action="store_true",
        help="Create timestamped backup before modifying file",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Process queries but don't write to Excel file",
    )

    return parser.parse_args()


async def main():
    """Main script execution."""
    args = parse_arguments()

    # Validate file exists
    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"✗ Error: File not found: {args.excel_file}")
        sys.exit(1)

    if not excel_path.suffix.lower() in [".xlsx", ".xlsm"]:
        print(f"✗ Error: File must be Excel format (.xlsx or .xlsm): {args.excel_file}")
        sys.exit(1)

    if args.start_sheet < 1:
        print(f"✗ Error: --start-sheet must be at least 1 (got {args.start_sheet})")
        sys.exit(1)

    # Load Excel file
    print(f"\n{'='*60}")
    print("Multi-Sheet Excel-to-LightRAG Query Processor")
    print(f"{'='*60}\n")

    processor = MultiSheetExcelProcessor(
        args.excel_file,
        prompt_column=args.prompt_column,
        output_column=args.output_column,
        start_sheet=args.start_sheet,
    )

    try:
        processor.load()
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        sys.exit(1)

    # Get sheets to process
    sheets_to_process = processor.get_sheets_to_process()
    print(f"✓ Sheets to process: {len(sheets_to_process)} (starting from sheet {args.start_sheet})")
    for ws in sheets_to_process:
        print(f"   - {ws.title}")

    print(f"✓ Reading prompts from column {args.prompt_column}")
    print(f"✓ Writing responses to column {args.output_column}")
    print(f"✓ API endpoint: {args.api_url}")
    print(f"✓ Max concurrent requests: {args.max_concurrent}")

    if args.dry_run:
        print("\n⚠ DRY RUN MODE - No changes will be written to Excel\n")

    # Handle authentication
    token = None
    if args.username and args.password:
        print(f"✓ Authenticating as: {args.username}")
        try:
            async with LightRAGClient(args.api_url, timeout=args.timeout) as auth_client:
                token = await auth_client.login(args.username, args.password)
            print("✓ Authentication successful")
        except aiohttp.ClientResponseError as e:
            if e.status == 401:
                print(f"\n✗ Error: Authentication failed - invalid username or password")
            else:
                print(f"\n✗ Error: Authentication failed - {e}")
            sys.exit(1)
        except aiohttp.ClientConnectorError:
            print(f"\n✗ Error: Cannot connect to LightRAG server at {args.api_url}")
            sys.exit(1)

    # Process each sheet
    total_success = 0
    total_errors = 0
    sheets_processed = 0

    try:
        async with LightRAGClient(
            args.api_url,
            api_key=args.api_key,
            token=token,
            timeout=args.timeout,
        ) as client:
            query_processor = ParallelQueryProcessor(client, args.max_concurrent)

            for worksheet in sheets_to_process:
                sheet_name = worksheet.title
                print(f"\n{'-'*60}")
                print(f"Processing sheet: {sheet_name}")
                print(f"{'-'*60}")

                # Extract prompts for this sheet
                prompts = processor.get_prompts(worksheet)
                if not prompts:
                    print(f"⚠ No prompts found in column {args.prompt_column} - skipping")
                    continue

                print(f"✓ Found {len(prompts)} prompts")

                # Process queries for this sheet
                results = await query_processor.process_all(prompts, sheet_name)

                # Write results to this sheet
                sheet_success = 0
                sheet_errors = 0

                for row_num, response, success in results:
                    if not args.dry_run:
                        processor.write_response(worksheet, row_num, response)
                    if success:
                        sheet_success += 1
                    else:
                        sheet_errors += 1

                total_success += sheet_success
                total_errors += sheet_errors
                sheets_processed += 1

                print(f"✓ Sheet '{sheet_name}': {sheet_success} success, {sheet_errors} errors")

    except aiohttp.ClientConnectorError:
        print(f"\n✗ Error: Cannot connect to LightRAG server at {args.api_url}")
        print("  Make sure the server is running with: lightrag-server")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error processing queries: {e}")
        sys.exit(1)

    # Save file (once at the end)
    if not args.dry_run and sheets_processed > 0:
        print(f"\n{'-'*60}")
        print("Saving results...")
        try:
            processor.save(backup=args.backup)
        except Exception as e:
            print(f"\n✗ Error saving Excel file: {e}")
            sys.exit(1)

    # Final summary
    print(f"\n{'='*60}")
    if args.dry_run:
        print(f"✓ Dry run complete!")
    else:
        print(f"✓ Processing complete!")
    print(f"  Sheets processed: {sheets_processed}")
    print(f"  Total success: {total_success}")
    print(f"  Total errors: {total_errors}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n✗ Interrupted by user")
        sys.exit(130)
