#!/usr/bin/env python3
"""
Excel-to-LightRAG Query Processor

This script processes an Excel file by:
1. Reading prompts from column B (header: "Pytanie")
2. Querying the LightRAG API in parallel (max 3 concurrent requests)
3. Writing responses to column U

Usage:
    python scripts/excel_lightrag_query.py --excel-file data.xlsx --api-url http://localhost:9621
"""

import argparse
import asyncio
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import aiohttp
import openpyxl
from openpyxl import load_workbook
from tenacity import (
    retry,
    stop_after_attempt,
    wait_exponential,
    retry_if_exception_type,
)

try:
    from tqdm.asyncio import tqdm as async_tqdm
except ImportError:
    print("Warning: tqdm not installed. Install with: pip install tqdm")
    print("Progress bar will not be available.")
    async_tqdm = None


class ExcelProcessor:
    """Handles Excel file reading and writing operations."""

    def __init__(self, file_path: str, prompt_column: str = "B", output_column: str = "U"):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.PROMPT_COLUMN = prompt_column.upper()
        self.OUTPUT_COLUMN = output_column.upper()

    def load(self):
        """Load Excel file and get active worksheet."""
        try:
            self.workbook = load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            print(f"✓ Loaded Excel file: {self.file_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file: {e}")

    def get_prompts(self) -> list[tuple[int, str]]:
        """
        Extract all non-empty prompts from the prompt column.
        Skips the header row (row 1).

        Returns:
            List of tuples (row_number, prompt_text)
        """
        prompts = []
        max_row = self.worksheet.max_row

        for row_num in range(2, max_row + 1):  # Skip header row
            cell_value = self.worksheet[f"{self.PROMPT_COLUMN}{row_num}"].value
            if cell_value and str(cell_value).strip():
                prompts.append((row_num, str(cell_value).strip()))

        return prompts

    def write_response(self, row_num: int, response: str):
        """
        Write response to the output column for the given row.
        Always overwrites existing data.

        Args:
            row_num: Excel row number
            response: Response text to write
        """
        self.worksheet[f"{self.OUTPUT_COLUMN}{row_num}"] = response

    def save(self, backup: bool = True):
        """
        Save the workbook with optional timestamped backup.

        Args:
            backup: If True, creates a backup before saving
        """
        if backup:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{self.file_path}.backup.{timestamp}"
            shutil.copy2(self.file_path, backup_path)
            print(f"✓ Backup created: {backup_path}")

        self.workbook.save(self.file_path)
        print(f"✓ Results saved to: {self.file_path}")


class LightRAGClient:
    """Async HTTP client for LightRAG API."""

    def __init__(
        self,
        base_url: str,
        api_key: Optional[str] = None,
        token: Optional[str] = None,
        timeout: int = 120,
    ):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.token = token
        self.timeout = aiohttp.ClientTimeout(total=timeout)
        self.session: Optional[aiohttp.ClientSession] = None

    async def __aenter__(self):
        """Create aiohttp session with timeout configuration."""
        self.session = aiohttp.ClientSession(timeout=self.timeout)
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Close aiohttp session."""
        if self.session:
            await self.session.close()

    async def login(self, username: str, password: str) -> str:
        """
        Authenticate with username/password and return JWT token.

        Args:
            username: Username for authentication
            password: Password for authentication

        Returns:
            JWT access token

        Raises:
            aiohttp.ClientError: On HTTP errors
        """
        url = f"{self.base_url}/login"
        data = aiohttp.FormData()
        data.add_field("username", username)
        data.add_field("password", password)

        async with self.session.post(url, data=data) as resp:
            resp.raise_for_status()
            result = await resp.json()
            return result.get("access_token")

    async def query(self, prompt: str) -> dict:
        """
        Send query to LightRAG API.

        Args:
            prompt: Query text

        Returns:
            JSON response from API

        Raises:
            aiohttp.ClientError: On HTTP errors
            asyncio.TimeoutError: On timeout
        """
        url = f"{self.base_url}/query"
        headers = {"Content-Type": "application/json"}

        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        elif self.api_key:
            headers["X-API-Key"] = self.api_key

        payload = {
            "query": prompt,
            "mode": "global",
            "top_k": 40,
            "chunk_top_k": 20,
            "max_entity_tokens": 6000,
            "max_relation_tokens": 8000,
            "max_total_tokens": 30000,
            "enable_rerank": False,
            "include_references": True,
            "response_type": "Single Paragraph"
        }

        async with self.session.post(url, json=payload, headers=headers) as resp:
            resp.raise_for_status()
            return await resp.json()


class ParallelQueryProcessor:
    """Manages parallel query processing with concurrency control."""

    def __init__(self, client: LightRAGClient, max_concurrent: int = 3):
        self.client = client
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.max_concurrent = max_concurrent

    async def process_query(
        self, row_num: int, prompt: str
    ) -> tuple[int, str, bool]:
        """
        Process a single query with retry logic.

        Args:
            row_num: Excel row number
            prompt: Query text

        Returns:
            Tuple of (row_number, response_text, success_flag)
        """
        async with self.semaphore:

            @retry(
                stop=stop_after_attempt(3),
                wait=wait_exponential(multiplier=1, min=2, max=10),
                retry=retry_if_exception_type(
                    (aiohttp.ClientError, asyncio.TimeoutError)
                ),
            )
            async def _query_with_retry():
                result = await self.client.query(prompt)
                return result.get("response", "")

            try:
                response_text = await _query_with_retry()
                return (row_num, response_text, True)
            except Exception as e:
                error_msg = f"ERROR (after 3 retries): {str(e)}"
                return (row_num, error_msg, False)

    async def process_all(
        self, prompts: list[tuple[int, str]]
    ) -> list[tuple[int, str, bool]]:
        """
        Process all prompts with progress tracking.

        Args:
            prompts: List of (row_number, prompt_text) tuples

        Returns:
            List of (row_number, response_text, success_flag) tuples
        """
        tasks = [
            self.process_query(row_num, prompt) for row_num, prompt in prompts
        ]

        if async_tqdm:
            # Use tqdm for detailed progress bar
            results = []
            for coro in async_tqdm.as_completed(
                tasks,
                desc="Processing queries",
                total=len(tasks),
                unit="query",
            ):
                results.append(await coro)
            # Sort by row number to maintain order
            results.sort(key=lambda x: x[0])
        else:
            # Fallback without progress bar
            print(f"Processing {len(tasks)} queries (max {self.max_concurrent} concurrent)...")
            results = await asyncio.gather(*tasks)

        return results


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Process Excel file with LightRAG queries",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage
  python scripts/excel_lightrag_query.py --excel-file data.xlsx

  # With authentication and backup
  python scripts/excel_lightrag_query.py --excel-file data.xlsx --api-key YOUR_KEY --backup

  # Custom configuration
  python scripts/excel_lightrag_query.py --excel-file data.xlsx --max-concurrent 5 --output-column V

Environment Variables:
  LIGHTRAG_API_URL      Default API URL (default: http://localhost:9621)
  LIGHTRAG_API_KEY      API key for authentication
  LIGHTRAG_USERNAME     Username for login authentication
  LIGHTRAG_PASSWORD     Password for login authentication
        """,
    )

    parser.add_argument(
        "--excel-file",
        required=True,
        help="Path to Excel file (.xlsx)",
    )

    parser.add_argument(
        "--api-url",
        default=os.getenv("LIGHTRAG_API_URL", "http://localhost:9621"),
        help="LightRAG API base URL (default: $LIGHTRAG_API_URL or http://localhost:9621)",
    )

    parser.add_argument(
        "--api-key",
        default=os.getenv("LIGHTRAG_API_KEY"),
        help="API key for authentication (default: $LIGHTRAG_API_KEY)",
    )

    parser.add_argument(
        "--username",
        default=os.getenv("LIGHTRAG_USERNAME"),
        help="Username for login authentication (default: $LIGHTRAG_USERNAME)",
    )

    parser.add_argument(
        "--password",
        default=os.getenv("LIGHTRAG_PASSWORD"),
        help="Password for login authentication (default: $LIGHTRAG_PASSWORD)",
    )

    parser.add_argument(
        "--max-concurrent",
        type=int,
        default=3,
        help="Maximum concurrent requests (default: 3)",
    )

    parser.add_argument(
        "--timeout",
        type=int,
        default=120,
        help="Request timeout in seconds (default: 120)",
    )

    parser.add_argument(
        "--output-column",
        default="U",
        help="Column for writing responses (default: U)",
    )

    parser.add_argument(
        "--prompt-column",
        default="B",
        help="Column containing prompts (default: B)",
    )

    parser.add_argument(
        "--backup",
        action="store_true",
        help="Create timestamped backup before modifying file",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Process queries but don't write to Excel file",
    )

    return parser.parse_args()


async def main():
    """Main script execution."""
    args = parse_arguments()

    # Validate file exists
    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"✗ Error: File not found: {args.excel_file}")
        sys.exit(1)

    if not excel_path.suffix.lower() in [".xlsx", ".xlsm"]:
        print(f"✗ Error: File must be Excel format (.xlsx or .xlsm): {args.excel_file}")
        sys.exit(1)

    # Load Excel file
    print(f"\n{'='*60}")
    print("Excel-to-LightRAG Query Processor")
    print(f"{'='*60}\n")

    processor = ExcelProcessor(
        args.excel_file,
        prompt_column=args.prompt_column,
        output_column=args.output_column,
    )

    try:
        processor.load()
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        sys.exit(1)

    # Extract prompts
    prompts = processor.get_prompts()
    if not prompts:
        print(f"✗ No prompts found in column {args.prompt_column}")
        sys.exit(1)

    print(f"✓ Found {len(prompts)} prompts in column {args.prompt_column}")
    print(f"✓ Will write responses to column {args.output_column}")
    print(f"✓ API endpoint: {args.api_url}")
    print(f"✓ Max concurrent requests: {args.max_concurrent}")

    if args.dry_run:
        print("\n⚠ DRY RUN MODE - No changes will be written to Excel\n")

    # Handle authentication
    token = None
    if args.username and args.password:
        print(f"✓ Authenticating as: {args.username}")
        try:
            async with LightRAGClient(args.api_url, timeout=args.timeout) as auth_client:
                token = await auth_client.login(args.username, args.password)
            print("✓ Authentication successful")
        except aiohttp.ClientResponseError as e:
            if e.status == 401:
                print(f"\n✗ Error: Authentication failed - invalid username or password")
            else:
                print(f"\n✗ Error: Authentication failed - {e}")
            sys.exit(1)
        except aiohttp.ClientConnectorError:
            print(f"\n✗ Error: Cannot connect to LightRAG server at {args.api_url}")
            sys.exit(1)

    # Process queries
    try:
        async with LightRAGClient(
            args.api_url,
            api_key=args.api_key,
            token=token,
            timeout=args.timeout,
        ) as client:
            query_processor = ParallelQueryProcessor(client, args.max_concurrent)
            print(f"\nProcessing queries...\n")
            results = await query_processor.process_all(prompts)
    except aiohttp.ClientConnectorError:
        print(f"\n✗ Error: Cannot connect to LightRAG server at {args.api_url}")
        print("  Make sure the server is running with: lightrag-server")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error processing queries: {e}")
        sys.exit(1)

    # Write results back to Excel
    if not args.dry_run:
        print("\n✓ Writing results to Excel...")
        success_count = 0
        error_count = 0

        for row_num, response, success in results:
            processor.write_response(row_num, response)
            if success:
                success_count += 1
            else:
                error_count += 1

        # Save file
        try:
            processor.save(backup=args.backup)
        except Exception as e:
            print(f"\n✗ Error saving Excel file: {e}")
            sys.exit(1)

        print(f"\n{'='*60}")
        print(f"✓ Complete! Success: {success_count}, Errors: {error_count}")
        print(f"{'='*60}\n")
    else:
        # Dry run - just show statistics
        success_count = sum(1 for _, _, success in results if success)
        error_count = sum(1 for _, _, success in results if not success)

        print(f"\n{'='*60}")
        print(f"✓ Dry run complete! Success: {success_count}, Errors: {error_count}")
        print(f"{'='*60}\n")

        if error_count > 0:
            print("Errors encountered:")
            for row_num, response, success in results:
                if not success:
                    print(f"  Row {row_num}: {response}")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n✗ Interrupted by user")
        sys.exit(130)
